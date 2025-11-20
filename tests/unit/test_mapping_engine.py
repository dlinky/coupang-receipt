"""Unit tests for mapping engine."""

import pytest
from openpyxl import Workbook
from src.services.mapping_engine import MappingEngine
from src.models.mapping_config import MappingConfiguration
from src.models.file_info import FileInformation
from src.services.config_manager import ConfigManager


def test_execute_simple_sum_adjacent_columns():
    """Test simple_sum with adjacent columns (G17:I46)."""
    # Create test workbooks
    head_workbook = Workbook()
    head_sheet = head_workbook.active
    head_sheet.title = "종합"
    
    # Fill test data: G17=10, H17=20, I17=30 (sum=60)
    #                 G18=5, H18=15, I18=25 (sum=45)
    head_sheet["G17"] = 10
    head_sheet["H17"] = 20
    head_sheet["I17"] = 30
    head_sheet["G18"] = 5
    head_sheet["H18"] = 15
    head_sheet["I18"] = 25
    
    branch_workbook = Workbook()
    branch_sheet = branch_workbook.create_sheet("주간정산")
    
    # Create mapping config
    mapping_config = MappingConfiguration(
        data_name="배달료",
        branch_sheet="주간정산",
        branch_range="E6:E7",
        head_office_sheet="종합",
        head_office_range="G17:I18",
        calculation_method="simple_sum"
    )
    
    # Execute mapping
    engine = MappingEngine()
    result = engine._execute_simple_sum(
        mapping_config, head_workbook, branch_workbook, week_offset=0
    )
    
    # Verify result
    assert result.success is True
    assert result.rows_processed == 2
    assert branch_sheet["E6"].value == 60  # 10+20+30
    assert branch_sheet["E7"].value == 45  # 5+15+25


def test_execute_simple_sum_non_adjacent_columns():
    """Test simple_sum with non-adjacent columns (K17:K18, M17:M18, O17:O18)."""
    # Create test workbooks
    head_workbook = Workbook()
    head_sheet = head_workbook.active
    head_sheet.title = "종합"
    
    # Fill test data
    head_sheet["K17"] = 10
    head_sheet["M17"] = 20
    head_sheet["O17"] = 30
    head_sheet["Q17"] = 40
    head_sheet["K18"] = 5
    head_sheet["M18"] = 15
    head_sheet["O18"] = 25
    head_sheet["Q18"] = 35
    
    branch_workbook = Workbook()
    branch_sheet = branch_workbook.create_sheet("주간정산")
    
    # Create mapping config
    mapping_config = MappingConfiguration(
        data_name="추가배달료",
        branch_sheet="주간정산",
        branch_range="E6:E7",
        head_office_sheet="종합",
        head_office_range="K17:K18, M17:M18, O17:O18, Q17:Q18",
        calculation_method="simple_sum"
    )
    
    # Execute mapping
    engine = MappingEngine()
    result = engine._execute_simple_sum(
        mapping_config, head_workbook, branch_workbook, week_offset=0
    )
    
    # Verify result
    assert result.success is True
    assert result.rows_processed == 2
    assert branch_sheet["E6"].value == 100  # 10+20+30+40
    assert branch_sheet["E7"].value == 80  # 5+15+25+35


def test_execute_simple_sum_with_empty_values():
    """Test simple_sum with None/empty values (should be ignored)."""
    # Create test workbooks
    head_workbook = Workbook()
    head_sheet = head_workbook.active
    head_sheet.title = "종합"
    
    # Fill test data with None values
    head_sheet["G17"] = 10
    head_sheet["H17"] = None
    head_sheet["I17"] = 20
    head_sheet["G18"] = ""
    head_sheet["H18"] = 15
    head_sheet["I18"] = 25
    
    branch_workbook = Workbook()
    branch_sheet = branch_workbook.create_sheet("주간정산")
    
    # Create mapping config
    mapping_config = MappingConfiguration(
        data_name="배달료",
        branch_sheet="주간정산",
        branch_range="E6:E7",
        head_office_sheet="종합",
        head_office_range="G17:I18",
        calculation_method="simple_sum"
    )
    
    # Execute mapping
    engine = MappingEngine()
    result = engine._execute_simple_sum(
        mapping_config, head_workbook, branch_workbook, week_offset=0
    )
    
    # Verify result (None and empty string should be ignored)
    assert result.success is True
    assert result.rows_processed == 2
    assert branch_sheet["E6"].value == 30  # 10+0+20 (None ignored)
    assert branch_sheet["E7"].value == 40  # 0+15+25 (empty string ignored)


def test_execute_simple_sum_with_week_offset():
    """Test simple_sum with week offset applied to branch file."""
    # Create test workbooks
    head_workbook = Workbook()
    head_sheet = head_workbook.active
    head_sheet.title = "종합"
    head_sheet["G17"] = 10
    head_sheet["H17"] = 20
    head_sheet["I17"] = 30
    
    branch_workbook = Workbook()
    branch_sheet = branch_workbook.create_sheet("주간정산")
    
    # Create mapping config
    mapping_config = MappingConfiguration(
        data_name="배달료",
        branch_sheet="주간정산",
        branch_range="E6:E6",  # Will be offset to E42
        head_office_sheet="종합",
        head_office_range="G17:I17",
        calculation_method="simple_sum"
    )
    
    # Execute mapping with offset (2주차 = +36)
    engine = MappingEngine()
    result = engine._execute_simple_sum(
        mapping_config, head_workbook, branch_workbook, week_offset=36
    )
    
    # Verify result (offset applied to target range)
    assert result.success is True
    assert result.rows_processed == 1
    assert branch_sheet["E42"].value == 60  # 10+20+30, written to E42 (E6+36)


def test_execute_simple_sum_mismatched_row_count():
    """Test simple_sum with mismatched row counts (should fail)."""
    # Create test workbooks
    head_workbook = Workbook()
    head_sheet = head_workbook.active
    head_sheet.title = "종합"
    head_sheet["G17"] = 10
    head_sheet["H17"] = 20
    head_sheet["G18"] = 5
    head_sheet["H18"] = 15
    
    branch_workbook = Workbook()
    branch_sheet = branch_workbook.create_sheet("주간정산")
    
    # Create mapping config with mismatched ranges
    mapping_config = MappingConfiguration(
        data_name="배달료",
        branch_sheet="주간정산",
        branch_range="E6:E7",
        head_office_sheet="종합",
        head_office_range="G17:H18, K17:K17",  # First has 2 rows, second has 1 row
        calculation_method="simple_sum"
    )
    
    # Execute mapping
    engine = MappingEngine()
    result = engine._execute_simple_sum(
        mapping_config, head_workbook, branch_workbook, week_offset=0
    )
    
    # Verify result (should fail)
    assert result.success is False
    assert "same number of rows" in result.error_message.lower()


def test_execute_conditional_sum_single_match():
    """Test conditional_sum with single matching row per rider."""
    # Create test workbooks
    head_workbook = Workbook()
    head_sheet = head_workbook.active
    head_sheet.title = "협력사 자체 미션"
    
    # Fill test data: row 9
    head_sheet["G9"] = "라이더1"  # name_column
    head_sheet["F9"] = 100  # value_column
    head_sheet["J9"] = "달성"  # check_column, check_value
    
    # row 10
    head_sheet["G10"] = "라이더2"
    head_sheet["F10"] = 200
    head_sheet["J10"] = "달성"
    
    branch_workbook = Workbook()
    branch_sheet = branch_workbook.create_sheet("주간정산")
    branch_sheet["C6"] = "라이더1"
    branch_sheet["C7"] = "라이더2"
    branch_sheet["C8"] = "라이더3"  # No match
    
    # Create mapping config
    from src.models.mapping_config import MappingConfiguration, Condition
    condition = Condition(
        source_sheet="협력사 자체 미션",
        name_column="G",
        value_column="F",
        check_column="J",
        check_value="달성"
    )
    mapping_config = MappingConfiguration(
        data_name="지점 프로모션 합계",
        branch_sheet="주간정산",
        branch_range="H6:H8",
        head_office_sheet="협력사 자체 미션",
        head_office_range="",
        calculation_method="conditional_sum",
        condition=condition
    )
    
    # Execute mapping
    engine = MappingEngine()
    result = engine._execute_conditional_sum(
        mapping_config, head_workbook, branch_workbook, week_offset=0
    )
    
    # Verify result
    assert result.success is True
    assert result.rows_processed == 2
    assert branch_sheet["H6"].value == 100  # 라이더1
    assert branch_sheet["H7"].value == 200  # 라이더2
    assert branch_sheet["H8"].value == 0  # 라이더3 (no match)


def test_execute_conditional_sum_multiple_matches():
    """Test conditional_sum with multiple matching rows for same rider."""
    # Create test workbooks
    head_workbook = Workbook()
    head_sheet = head_workbook.active
    head_sheet.title = "협력사 자체 미션"
    
    # Fill test data: 라이더1이 여러 행에 있음
    head_sheet["G9"] = "라이더1"
    head_sheet["F9"] = 100
    head_sheet["J9"] = "달성"
    
    head_sheet["G10"] = "라이더1"  # Same rider
    head_sheet["F10"] = 150
    head_sheet["J10"] = "달성"
    
    head_sheet["G11"] = "라이더1"  # Same rider again
    head_sheet["F11"] = 50
    head_sheet["J11"] = "달성"
    
    head_sheet["G12"] = "라이더2"
    head_sheet["F12"] = 200
    head_sheet["J12"] = "달성"
    
    branch_workbook = Workbook()
    branch_sheet = branch_workbook.create_sheet("주간정산")
    branch_sheet["C6"] = "라이더1"
    branch_sheet["C7"] = "라이더2"
    
    # Create mapping config
    from src.models.mapping_config import MappingConfiguration, Condition
    condition = Condition(
        source_sheet="협력사 자체 미션",
        name_column="G",
        value_column="F",
        check_column="J",
        check_value="달성"
    )
    mapping_config = MappingConfiguration(
        data_name="지점 프로모션 합계",
        branch_sheet="주간정산",
        branch_range="H6:H7",
        head_office_sheet="협력사 자체 미션",
        head_office_range="",
        calculation_method="conditional_sum",
        condition=condition
    )
    
    # Execute mapping
    engine = MappingEngine()
    result = engine._execute_conditional_sum(
        mapping_config, head_workbook, branch_workbook, week_offset=0
    )
    
    # Verify result (라이더1의 값들이 합산됨)
    assert result.success is True
    assert result.rows_processed == 2
    assert branch_sheet["H6"].value == 300  # 100 + 150 + 50
    assert branch_sheet["H7"].value == 200  # 라이더2


def test_execute_conditional_sum_with_non_matching_condition():
    """Test conditional_sum with rows that don't match condition."""
    # Create test workbooks
    head_workbook = Workbook()
    head_sheet = head_workbook.active
    head_sheet.title = "협력사 자체 미션"
    
    # Fill test data: check_value가 "달성"이 아닌 경우
    head_sheet["G9"] = "라이더1"
    head_sheet["F9"] = 100
    head_sheet["J9"] = "미달성"  # Not matching
    
    head_sheet["G10"] = "라이더1"
    head_sheet["F10"] = 200
    head_sheet["J10"] = "달성"  # Matching
    
    branch_workbook = Workbook()
    branch_sheet = branch_workbook.create_sheet("주간정산")
    branch_sheet["C6"] = "라이더1"
    
    # Create mapping config
    from src.models.mapping_config import MappingConfiguration, Condition
    condition = Condition(
        source_sheet="협력사 자체 미션",
        name_column="G",
        value_column="F",
        check_column="J",
        check_value="달성"
    )
    mapping_config = MappingConfiguration(
        data_name="지점 프로모션 합계",
        branch_sheet="주간정산",
        branch_range="H6:H6",
        head_office_sheet="협력사 자체 미션",
        head_office_range="",
        calculation_method="conditional_sum",
        condition=condition
    )
    
    # Execute mapping
    engine = MappingEngine()
    result = engine._execute_conditional_sum(
        mapping_config, head_workbook, branch_workbook, week_offset=0
    )
    
    # Verify result (only matching row is summed)
    assert result.success is True
    assert result.rows_processed == 1
    assert branch_sheet["H6"].value == 200  # Only F10 (200) is summed, F9 (100) is ignored


def test_execute_conditional_sum_with_non_numeric_values():
    """Test conditional_sum with non-numeric values (should be ignored)."""
    # Create test workbooks
    head_workbook = Workbook()
    head_sheet = head_workbook.active
    head_sheet.title = "협력사 자체 미션"
    
    # Fill test data with None and non-numeric values
    head_sheet["G9"] = "라이더1"
    head_sheet["F9"] = 100
    head_sheet["J9"] = "달성"
    
    head_sheet["G10"] = "라이더1"
    head_sheet["F10"] = None  # Should be ignored
    head_sheet["J10"] = "달성"
    
    head_sheet["G11"] = "라이더1"
    head_sheet["F11"] = "문자열"  # Should be ignored
    head_sheet["J11"] = "달성"
    
    head_sheet["G12"] = "라이더1"
    head_sheet["F12"] = 50
    head_sheet["J12"] = "달성"
    
    branch_workbook = Workbook()
    branch_sheet = branch_workbook.create_sheet("주간정산")
    branch_sheet["C6"] = "라이더1"
    
    # Create mapping config
    from src.models.mapping_config import MappingConfiguration, Condition
    condition = Condition(
        source_sheet="협력사 자체 미션",
        name_column="G",
        value_column="F",
        check_column="J",
        check_value="달성"
    )
    mapping_config = MappingConfiguration(
        data_name="지점 프로모션 합계",
        branch_sheet="주간정산",
        branch_range="H6:H6",
        head_office_sheet="협력사 자체 미션",
        head_office_range="",
        calculation_method="conditional_sum",
        condition=condition
    )
    
    # Execute mapping
    engine = MappingEngine()
    result = engine._execute_conditional_sum(
        mapping_config, head_workbook, branch_workbook, week_offset=0
    )
    
    # Verify result (only numeric values are summed)
    assert result.success is True
    assert result.rows_processed == 1
    assert branch_sheet["H6"].value == 150  # 100 + 50 (None and string ignored)


def test_execute_conditional_sum_with_week_offset():
    """Test conditional_sum with week offset applied to branch file."""
    # Create test workbooks
    head_workbook = Workbook()
    head_sheet = head_workbook.active
    head_sheet.title = "협력사 자체 미션"
    head_sheet["G9"] = "라이더1"
    head_sheet["F9"] = 100
    head_sheet["J9"] = "달성"
    
    branch_workbook = Workbook()
    branch_sheet = branch_workbook.create_sheet("주간정산")
    branch_sheet["C6"] = "라이더1"
    
    # Create mapping config
    from src.models.mapping_config import MappingConfiguration, Condition
    condition = Condition(
        source_sheet="협력사 자체 미션",
        name_column="G",
        value_column="F",
        check_column="J",
        check_value="달성"
    )
    mapping_config = MappingConfiguration(
        data_name="지점 프로모션 합계",
        branch_sheet="주간정산",
        branch_range="H6:H6",  # Will be offset to H42
        head_office_sheet="협력사 자체 미션",
        head_office_range="",
        calculation_method="conditional_sum",
        condition=condition
    )
    
    # Execute mapping with offset (2주차 = +36)
    engine = MappingEngine()
    result = engine._execute_conditional_sum(
        mapping_config, head_workbook, branch_workbook, week_offset=36
    )
    
    # Verify result (offset applied to target range)
    assert result.success is True
    assert result.rows_processed == 1
    assert branch_sheet["H42"].value == 100  # Written to H42 (H6+36)

