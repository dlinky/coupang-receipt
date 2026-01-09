"""Mapping engine service."""

from dataclasses import dataclass
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from ..models.mapping_config import MappingConfiguration
from ..models.file_info import FileInformation
from ..exceptions import MappingError, FileError
from .excel_processor import ExcelProcessor
from .config_manager import ConfigManager
from .date_calculator import DateCalculator
from ..utils.cell_utils import CellUtils
from ..utils.file_parser import FileParser
from pathlib import Path


@dataclass
class MappingResult:
    """Result of mapping operation."""
    success: bool
    rows_processed: int
    error_message: Optional[str] = None


class MappingEngine:
    """Core mapping engine for data transformation."""
    
    def __init__(self, config_manager: Optional[ConfigManager] = None):
        """Initialize mapping engine."""
        self.excel_processor = ExcelProcessor()
        self.config_manager = config_manager or ConfigManager()
        self.date_calculator = DateCalculator()
        self._mappings: List[MappingConfiguration] = []
    
    def load_mappings(self) -> List[MappingConfiguration]:
        """Load mapping configurations."""
        self._mappings = self.config_manager.load_mappings()
        return self._mappings
    
    def execute_mapping(
        self,
        mapping_config: MappingConfiguration,
        head_office_file: FileInformation,
        branch_file_path: str,
        week: int,
        branch_workbook: Optional[Workbook] = None
    ) -> MappingResult:
        """Execute mapping operation.
        
        Args:
            mapping_config: Mapping configuration
            head_office_file: Head office file information
            branch_file_path: Branch file path
            week: Week number (1-8)
            branch_workbook: Optional workbook object (if provided, will be reused)
        
        Returns:
            MappingResult with success status and details
        """
        if not (1 <= week <= 8):
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"Week must be between 1 and 8, got {week}"
            )
        
        try:
            head_workbook = self.excel_processor.load_workbook(
                head_office_file.file_path,
                head_office_file.password
            )
            
            if branch_workbook is None:
                branch_workbook = self.excel_processor.load_workbook(branch_file_path, data_only=False)
            
            # Use the manually selected week for offset calculation, not the file's original week
            week_offset = self.config_manager.get_week_offset(week)
            
            if mapping_config.calculation_method == "simple_copy":
                result = self._execute_simple_copy(
                    mapping_config, head_workbook, branch_workbook, week_offset
                )
            elif mapping_config.calculation_method == "conditional_copy":
                result = self._execute_conditional_copy(
                    mapping_config, head_workbook, branch_workbook, week_offset
                )
            elif mapping_config.calculation_method == "date_calculation":
                result = self._execute_date_calculation(
                    mapping_config, branch_workbook, head_office_file, week, week_offset
                )
            elif mapping_config.calculation_method == "unique_extraction":
                result = self._execute_unique_extraction(
                    mapping_config, branch_workbook
                )
            elif mapping_config.calculation_method == "simple_sum":
                result = self._execute_simple_sum(
                    mapping_config, head_workbook, branch_workbook, week_offset
                )
            elif mapping_config.calculation_method == "conditional_sum":
                result = self._execute_conditional_sum(
                    mapping_config, head_workbook, branch_workbook, week_offset
                )
            elif mapping_config.calculation_method == "fee":
                result = self._execute_fee(
                    mapping_config, branch_workbook, week_offset
                )
            else:
                return MappingResult(
                    success=False,
                    rows_processed=0,
                    error_message=f"Unknown calculation method: {mapping_config.calculation_method}"
                )
            
            return result
            
        except FileNotFoundError as e:
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"File not found: {e}"
            )
        except PermissionError as e:
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"Permission error: {e}"
            )
        except ValueError as e:
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"Invalid mapping configuration: {e}"
            )
        except KeyError as e:
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"Sheet not found: {e}"
            )
        except Exception as e:
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"Unexpected error: {e}"
            )
    
    def save_branch_file(
        self,
        branch_workbook: Workbook,
        branch_file_path: str,
        head_office_file: FileInformation,
        week: int
    ) -> str:
        """Save branch file with correct filename.
        
        Args:
            branch_workbook: Workbook to save
            branch_file_path: Original branch file path
            head_office_file: Head office file information
            week: Week number (1-8)
        
        Returns:
            New file path where file was saved
        """
        branch_file_path_obj = Path(branch_file_path)
        new_filename = FileParser.generate_branch_filename(
            head_office_file.year, head_office_file.month, week
        )
        new_file_path = branch_file_path_obj.parent / new_filename
        
        self.excel_processor.save_workbook(branch_workbook, str(new_file_path))
        
        return str(new_file_path)
    
    def _execute_simple_copy(
        self,
        mapping_config: MappingConfiguration,
        head_workbook: Workbook,
        branch_workbook: Workbook,
        week_offset: int
    ) -> MappingResult:
        """Execute simple copy operation."""
        try:
            head_sheet = self.excel_processor.get_sheet(
                head_workbook, mapping_config.head_office_sheet
            )
            branch_sheet = self.excel_processor.get_sheet(
                branch_workbook, mapping_config.branch_sheet
            )
            
            source_range = mapping_config.head_office_range
            target_range = CellUtils.apply_row_offset(
                mapping_config.branch_range, week_offset
            )
            
            values = self.excel_processor.read_cell_range(head_sheet, source_range)
            
            if not values:
                return MappingResult(success=True, rows_processed=0)
            
            if mapping_config.data_name in ["고용보험", "산재보험", "시간제보험"]:
                values = [v * -1 if isinstance(v, (int, float)) and v is not None else v for v in values]
            
            self.excel_processor.write_cell_range(branch_sheet, target_range, values)
            
            return MappingResult(success=True, rows_processed=len(values))
            
        except Exception as e:
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"Simple copy failed: {e}"
            )
    
    def _execute_simple_sum(
        self,
        mapping_config: MappingConfiguration,
        head_workbook: Workbook,
        branch_workbook: Workbook,
        week_offset: int
    ) -> MappingResult:
        """Execute simple sum operation - sum values from multiple columns.
        
        Args:
            mapping_config: Mapping configuration with head_office_range containing
                          comma-separated ranges (e.g., "G17:I46" or "K17:K46, M17:M46")
            head_workbook: Head office workbook
            branch_workbook: Branch workbook
            week_offset: Week offset for branch file
        
        Returns:
            MappingResult with success status and rows processed
        """
        try:
            head_sheet = self.excel_processor.get_sheet(
                head_workbook, mapping_config.head_office_sheet
            )
            branch_sheet = self.excel_processor.get_sheet(
                branch_workbook, mapping_config.branch_sheet
            )
            
            # Parse multiple ranges from head_office_range
            source_ranges = CellUtils.parse_multiple_ranges(mapping_config.head_office_range)
            
            # Parse target range and apply offset
            target_range = CellUtils.apply_row_offset(
                mapping_config.branch_range, week_offset
            )
            target_start_col, target_start_row, target_end_col, target_end_row = CellUtils.parse_range(target_range)
            
            # Determine number of rows from first source range
            first_range = source_ranges[0]
            first_start_col, first_start_row, first_end_col, first_end_row = CellUtils.parse_range(first_range)
            num_rows = first_end_row - first_start_row + 1
            
            # Validate that all ranges have the same number of rows
            for range_str in source_ranges[1:]:
                start_col, start_row, end_col, end_row = CellUtils.parse_range(range_str)
                if (end_row - start_row + 1) != num_rows:
                    return MappingResult(
                        success=False,
                        rows_processed=0,
                        error_message=f"All ranges must have the same number of rows. "
                                     f"First range has {num_rows} rows, but {range_str} has {end_row - start_row + 1} rows"
                    )
            
            # Read rider names from branch file (C column) to check if we should write values
            rider_column = "C"
            rider_range = f"{rider_column}{target_start_row}:{rider_column}{target_end_row}"
            rider_names = self.excel_processor.read_cell_range(branch_sheet, rider_range)
            
            # Read values from each range and sum row by row
            summed_values = []
            for row_idx in range(num_rows):
                # Check if rider name exists for this row
                rider_name = rider_names[row_idx] if row_idx < len(rider_names) else None
                if not rider_name or not str(rider_name).strip():
                    # No rider name, skip this row (set to None)
                    summed_values.append(None)
                    continue
                
                row_sum = 0
                current_row = first_start_row + row_idx
                
                # Sum values from all ranges for this row
                for range_str in source_ranges:
                    start_col, start_row, end_col, end_row = CellUtils.parse_range(range_str)
                    # Read all columns in this range for current row
                    start_col_idx = column_index_from_string(start_col)
                    end_col_idx = column_index_from_string(end_col)
                    
                    for col_idx in range(start_col_idx, end_col_idx + 1):
                        col_letter = get_column_letter(col_idx)
                        cell_address = f"{col_letter}{current_row}"
                        
                        # Handle merged cells
                        merged_cell = self.excel_processor.get_merged_cell_top_left(head_sheet, cell_address)
                        if merged_cell:
                            cell_address = merged_cell
                        
                        cell = head_sheet[cell_address]
                        cell_value = cell.value
                        
                        # Sum numeric values, ignore None and non-numeric
                        if isinstance(cell_value, (int, float)):
                            row_sum += cell_value
                
                summed_values.append(row_sum)
            
            # Write summed values to branch file
            if not summed_values:
                return MappingResult(success=True, rows_processed=0)
            
            self.excel_processor.write_cell_range(branch_sheet, target_range, summed_values)
            
            processed = len([v for v in summed_values if v is not None])
            return MappingResult(success=True, rows_processed=processed)
            
        except Exception as e:
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"Simple sum failed: {e}"
            )
    
    def _execute_conditional_copy(
        self,
        mapping_config: MappingConfiguration,
        head_workbook: Workbook,
        branch_workbook: Workbook,
        week_offset: int
    ) -> MappingResult:
        """Execute conditional copy operation."""
        try:
            if not mapping_config.condition:
                return MappingResult(
                    success=False,
                    rows_processed=0,
                    error_message="Condition is required for conditional copy"
                )
            
            condition = mapping_config.condition
            head_sheet = self.excel_processor.get_sheet(
                head_workbook, condition.source_sheet
            )
            branch_sheet = self.excel_processor.get_sheet(
                branch_workbook, mapping_config.branch_sheet
            )
            
            target_range = CellUtils.apply_row_offset(
                mapping_config.branch_range, week_offset
            )
            start_col, start_row, end_col, end_row = CellUtils.parse_range(target_range)
            
            rider_column = "C"
            rider_range = f"{rider_column}{start_row}:{rider_column}{end_row}"
            rider_names = self.excel_processor.read_cell_range(branch_sheet, rider_range)
            
            name_col_idx = column_index_from_string(condition.name_column)
            value_col_idx = column_index_from_string(condition.value_column)
            check_col_idx = column_index_from_string(condition.check_column)
            
            promotion_values = {}
            
            for row in range(9, head_sheet.max_row + 1):
                name_cell = head_sheet.cell(row, name_col_idx)
                value_cell = head_sheet.cell(row, value_col_idx)
                check_cell = head_sheet.cell(row, check_col_idx)
                
                if name_cell.value and check_cell.value == condition.check_value:
                    rider_name = str(name_cell.value).strip()
                    promotion_value = value_cell.value
                    promotion_values[rider_name] = promotion_value
            
            result_values = []
            for rider_name in rider_names:
                if rider_name and str(rider_name).strip() in promotion_values:
                    result_values.append(promotion_values[str(rider_name).strip()])
                else:
                    result_values.append(None)
            
            self.excel_processor.write_cell_range(branch_sheet, target_range, result_values)
            
            processed = len([v for v in result_values if v is not None])
            return MappingResult(success=True, rows_processed=processed)
            
        except Exception as e:
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"Conditional copy failed: {e}"
            )
    
    def _execute_conditional_sum(
        self,
        mapping_config: MappingConfiguration,
        head_workbook: Workbook,
        branch_workbook: Workbook,
        week_offset: int
    ) -> MappingResult:
        """Execute conditional sum operation - sum value_column values for rows matching condition.
        
        Similar to conditional_copy, but sums all matching values instead of taking the first one.
        
        Args:
            mapping_config: Mapping configuration with condition
            head_workbook: Head office workbook
            branch_workbook: Branch workbook
            week_offset: Week offset for branch file
        
        Returns:
            MappingResult with success status and rows processed
        """
        try:
            if not mapping_config.condition:
                return MappingResult(
                    success=False,
                    rows_processed=0,
                    error_message="Condition is required for conditional sum"
                )
            
            condition = mapping_config.condition
            head_sheet = self.excel_processor.get_sheet(
                head_workbook, condition.source_sheet
            )
            branch_sheet = self.excel_processor.get_sheet(
                branch_workbook, mapping_config.branch_sheet
            )
            
            target_range = CellUtils.apply_row_offset(
                mapping_config.branch_range, week_offset
            )
            start_col, start_row, end_col, end_row = CellUtils.parse_range(target_range)
            
            rider_column = "C"
            rider_range = f"{rider_column}{start_row}:{rider_column}{end_row}"
            rider_names = self.excel_processor.read_cell_range(branch_sheet, rider_range)
            
            name_col_idx = column_index_from_string(condition.name_column)
            value_col_idx = column_index_from_string(condition.value_column)
            check_col_idx = column_index_from_string(condition.check_column)
            
            # Sum values for each rider (multiple rows can match)
            promotion_sums = {}
            
            for row in range(9, head_sheet.max_row + 1):
                name_cell = head_sheet.cell(row, name_col_idx)
                value_cell = head_sheet.cell(row, value_col_idx)
                check_cell = head_sheet.cell(row, check_col_idx)
                
                if name_cell.value and check_cell.value == condition.check_value:
                    rider_name = str(name_cell.value).strip()
                    value = value_cell.value
                    
                    # Sum numeric values, ignore None and non-numeric
                    if isinstance(value, (int, float)):
                        if rider_name not in promotion_sums:
                            promotion_sums[rider_name] = 0
                        promotion_sums[rider_name] += value
            
            result_values = []
            for rider_name in rider_names:
                if rider_name and str(rider_name).strip():
                    rider_name_str = str(rider_name).strip()
                    if rider_name_str in promotion_sums:
                        result_values.append(promotion_sums[rider_name_str])
                    else:
                        result_values.append(None)  # No matching rows, don't write
                else:
                    result_values.append(None)  # Empty rider name, don't write
            
            self.excel_processor.write_cell_range(branch_sheet, target_range, result_values)
            
            processed = len([v for v in result_values if v is not None])
            return MappingResult(success=True, rows_processed=processed)
            
        except Exception as e:
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"Conditional sum failed: {e}"
            )
    
    def _execute_date_calculation(
        self,
        mapping_config: MappingConfiguration,
        branch_workbook: Workbook,
        head_office_file: FileInformation,
        week: int,
        week_offset: int = 0
    ) -> MappingResult:
        """Execute date calculation."""
        try:
            if not mapping_config.date_type:
                return MappingResult(
                    success=False,
                    rows_processed=0,
                    error_message="Date type is required for date calculation"
                )
            
            branch_sheet = self.excel_processor.get_sheet(
                branch_workbook, mapping_config.branch_sheet
            )
            
            date_str = ""
            is_monthly = mapping_config.branch_sheet == "월간정산"
            
            if mapping_config.date_type == "title":
                if is_monthly:
                    base_date = f"{head_office_file.month}월"
                else:
                    base_date = self.date_calculator.calculate_title_date(
                        head_office_file.year, head_office_file.month, week
                    )
                
                if mapping_config.data_name == "기사별 정산 내역 타이틀":
                    date_str = f"기사별 정산 내역({base_date})"
                elif mapping_config.data_name == "익일정산 신청 내역 타이틀":
                    date_str = f"익일정산 신청 내역({base_date})"
                elif mapping_config.data_name == "지점 정산서 타이틀":
                    date_str = f"지점 정산서({base_date})"
                else:
                    date_str = base_date
                    
            elif mapping_config.date_type == "payment_date":
                date_str = self.date_calculator.calculate_payment_date(
                    head_office_file.year, head_office_file.month, week
                )
            elif mapping_config.date_type == "date_range":
                date_range = self.date_calculator.calculate_date_range(
                    head_office_file.year, head_office_file.month, week
                )
                
                if mapping_config.data_name == "기사별 정산서 타이틀":
                    if is_monthly:
                        base_date = f"{head_office_file.month}월"
                    else:
                        base_date = self.date_calculator.calculate_title_date(
                            head_office_file.year, head_office_file.month, week
                        )
                    date_str = f"쿠팡 {base_date} 정산서\n({date_range})"
                else:
                    date_str = date_range
            else:
                return MappingResult(
                    success=False,
                    rows_processed=0,
                    error_message=f"Unknown date type: {mapping_config.date_type}"
                )
            
            target_range = mapping_config.branch_range
            if not is_monthly and week_offset != 0:
                target_range = CellUtils.apply_row_offset(mapping_config.branch_range, week_offset)
            
            self.excel_processor.write_cell_range(
                branch_sheet, target_range, [date_str]
            )
            
            return MappingResult(success=True, rows_processed=1)
            
        except Exception as e:
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"Date calculation failed: {e}"
            )
    
    def _execute_unique_extraction(
        self,
        mapping_config: MappingConfiguration,
        branch_workbook: Workbook
    ) -> MappingResult:
        """Execute unique extraction for monthly settlement.
        
        Extracts unique values from multiple columns across all weeks,
        sorts by specified column and order, and writes all column values to target range.
        """
        try:
            # Get source range from mapping config (e.g., "C6:E35")
            source_range_str = mapping_config.branch_range
            start_col, start_row, end_col, end_row = CellUtils.parse_range(source_range_str)
            
            # Get sort column from config (default to D if not specified)
            sort_col = mapping_config.sort_column or "D"
            sort_order = (mapping_config.sort_order or "asc").lower()
            
            # Validate sort column is within source range
            start_col_idx = column_index_from_string(start_col)
            end_col_idx = column_index_from_string(end_col)
            sort_col_idx = column_index_from_string(sort_col)
            
            if not (start_col_idx <= sort_col_idx <= end_col_idx):
                return MappingResult(
                    success=False,
                    rows_processed=0,
                    error_message=f"Sort column '{sort_col}' is not within source range '{source_range_str}'"
                )
            
            weekly_sheet = self.excel_processor.get_sheet(
                branch_workbook, "주간정산"
            )
            
            # Collect all rows from all weeks
            all_rows = []
            
            for week in range(1, 9):
                week_offset = self.config_manager.get_week_offset(week)
                # Apply week offset to source range
                source_range = CellUtils.apply_row_offset(source_range_str, week_offset)
                start_col_offset, start_row_offset, end_col_offset, end_row_offset = CellUtils.parse_range(source_range)
                
                # Read each row as a tuple of all column values
                for row in range(start_row_offset, end_row_offset + 1):
                    row_values = []
                    for col_idx in range(start_col_idx, end_col_idx + 1):
                        cell = weekly_sheet.cell(row, col_idx)
                        row_values.append(cell.value)
                    
                    # Skip empty rows (at least first column value should exist)
                    if row_values[0] and str(row_values[0]).strip():
                        all_rows.append(tuple(row_values))
            
            # Extract unique rows using set
            unique_rows = list(set(all_rows))
            
            # Get sort column index relative to start column
            sort_col_relative_idx = sort_col_idx - start_col_idx
            
            # Sort by specified column
            def sort_key(row):
                sort_val = row[sort_col_relative_idx] if sort_col_relative_idx < len(row) else None
                if sort_val is None:
                    return (1, "")  # Put None values at the end
                # Convert to string for consistent sorting
                sort_str = str(sort_val).strip() if sort_val else ""
                return (0, sort_str)
            
            unique_rows_sorted = sorted(unique_rows, key=sort_key, reverse=(sort_order == "desc"))
            
            monthly_sheet = self.excel_processor.get_sheet(
                branch_workbook, mapping_config.branch_sheet
            )
            
            # Parse target range
            target_range = mapping_config.branch_range
            target_start_col, target_start_row, target_end_col, target_end_row = CellUtils.parse_range(target_range)
            
            # Calculate number of columns and rows from target range
            target_start_col_idx = column_index_from_string(target_start_col)
            target_end_col_idx = column_index_from_string(target_end_col)
            num_cols = target_end_col_idx - target_start_col_idx + 1
            num_rows = target_end_row - target_start_row + 1
            
            # Flatten sorted rows into a list for writing (row by row, column by column)
            # Format: C6, D6, E6, C7, D7, E7, ..., C35, D35, E35
            values = []
            for i in range(num_rows):
                if i < len(unique_rows_sorted):
                    row_data = unique_rows_sorted[i]
                    # Ensure row_data has enough columns
                    for col_idx in range(num_cols):
                        if col_idx < len(row_data):
                            values.append(row_data[col_idx])
                        else:
                            values.append(None)
                else:
                    # Fill empty rows with None values
                    values.extend([None] * num_cols)
            
            self.excel_processor.write_cell_range(monthly_sheet, target_range, values)
            
            return MappingResult(success=True, rows_processed=len(unique_rows_sorted))
            
        except Exception as e:
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"Unique extraction failed: {e}"
            )
    
    def _execute_fee(
        self,
        mapping_config: MappingConfiguration,
        branch_workbook: Workbook,
        week_offset: int
    ) -> MappingResult:
        """Execute fee calculation operation.
        
        Writes fee formula (=F{row}*{fee_ratio}) only for checked riders.
        """
        try:
            if mapping_config.fee_ratio is None:
                return MappingResult(
                    success=False,
                    rows_processed=0,
                    error_message="fee_ratio is required for fee calculation"
                )
            
            branch_sheet = self.excel_processor.get_sheet(
                branch_workbook, mapping_config.branch_sheet
            )
            
            # Get fee riders from config
            fee_riders = self.config_manager.get_fee_riders()
            if not fee_riders:
                return MappingResult(success=True, rows_processed=0)
            
            # Find "성함" mapping to get rider names from branch file
            mappings = self.load_mappings()
            name_mapping = None
            for m in mappings:
                if m.data_name == "성함":
                    name_mapping = m
                    break
            
            if not name_mapping:
                return MappingResult(
                    success=False,
                    rows_processed=0,
                    error_message="'성함' mapping not found"
                )
            
            # Apply week offset to name range
            name_range = CellUtils.apply_row_offset(name_mapping.branch_range, week_offset)
            start_col, start_row, end_col, end_row = CellUtils.parse_range(name_range)
            
            # Read rider names from branch file
            rider_names = self.excel_processor.read_cell_range(branch_sheet, name_range)
            
            # Parse target range and apply offset
            target_range = CellUtils.apply_row_offset(
                mapping_config.branch_range, week_offset
            )
            target_start_col, target_start_row, target_end_col, target_end_row = CellUtils.parse_range(target_range)
            
            # Get fee base column (F column for delivery fee)
            # This could be made configurable later, but for now F is hardcoded
            fee_base_col = "F"
            
            processed_count = 0
            for i, rider_name in enumerate(rider_names):
                if not rider_name or not str(rider_name).strip():
                    continue
                
                rider_name_str = str(rider_name).strip()
                
                # Check if this rider has fee applied
                if rider_name_str in fee_riders:
                    # Calculate target row
                    target_row = target_start_row + i
                    
                    # Create formula: =F{row}*{fee_ratio}
                    fee_base_row = start_row + i  # F column row (same as name row)
                    formula = f"={fee_base_col}{fee_base_row}*{mapping_config.fee_ratio}"
                    
                    # Write formula to cell
                    cell_address = f"{target_start_col}{target_row}"
                    
                    # Handle merged cells
                    merged_cell = self.excel_processor.get_merged_cell_top_left(branch_sheet, cell_address)
                    if merged_cell:
                        cell_address = merged_cell
                    
                    cell = branch_sheet[cell_address]
                    cell.value = formula
                    processed_count += 1
            
            return MappingResult(success=True, rows_processed=processed_count)
            
        except Exception as e:
            return MappingResult(
                success=False,
                rows_processed=0,
                error_message=f"Fee calculation failed: {e}"
            )

