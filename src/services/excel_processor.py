"""Excel file processing service."""

from pathlib import Path
from typing import List, Any, Optional
import io
import msoffcrypto
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment

from ..exceptions import FileError
from ..utils.cell_utils import CellUtils


class ExcelProcessor:
    """Handles Excel file operations."""
    
    def load_workbook(self, file_path: str, password: Optional[str] = None, data_only: bool = False) -> Workbook:
        """Load Excel workbook.
        
        Args:
            file_path: Path to Excel file
            password: Password for protected files
            data_only: If True, only load calculated values (formulas are lost). 
                      If False, preserve formulas. Default False to preserve formulas.
        
        Returns:
            Workbook object
        
        Raises:
            FileNotFoundError: File does not exist
            ValueError: Invalid password
            PermissionError: No access permission
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        try:
            if password:
                decrypted = io.BytesIO()
                with open(path, "rb") as file:
                    office_file = msoffcrypto.OfficeFile(file)
                    office_file.load_key(password=password)
                    office_file.decrypt(decrypted)
                
                decrypted.seek(0)
                workbook = load_workbook(decrypted, read_only=False, keep_vba=False, data_only=data_only)
            else:
                workbook = load_workbook(path, read_only=False, keep_vba=False, data_only=data_only)
            return workbook
        except msoffcrypto.exceptions.InvalidKeyError:
            raise ValueError(f"Invalid password for file: {file_path}")
        except Exception as e:
            if "password" in str(e).lower() or "encrypted" in str(e).lower():
                raise ValueError(f"Invalid password for file: {file_path}") from e
            raise FileError(f"Failed to load workbook: {e}") from e
    
    def get_sheet(self, workbook: Workbook, sheet_name: str) -> Worksheet:
        """Get worksheet from workbook.
        
        Args:
            workbook: Workbook object
            sheet_name: Name of the sheet
        
        Returns:
            Worksheet object
        
        Raises:
            KeyError: Sheet does not exist
        """
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found in workbook")
        return workbook[sheet_name]
    
    def read_cell_range(self, worksheet: Worksheet, cell_range: str) -> List[Any]:
        """Read values from cell range.
        
        Args:
            worksheet: Worksheet object
            cell_range: Cell range like "A1:B10" or "C6:C35"
        
        Returns:
            List of cell values
        
        Raises:
            ValueError: Invalid cell range format
        """
        try:
            start_col, start_row, end_col, end_row = CellUtils.parse_range(cell_range)
            
            values = []
            for row in range(start_row, end_row + 1):
                row_values = []
                start_col_idx = column_index_from_string(start_col)
                end_col_idx = column_index_from_string(end_col)
                
                for col_idx in range(start_col_idx, end_col_idx + 1):
                    col_letter = get_column_letter(col_idx)
                    cell_address = f"{col_letter}{row}"
                    
                    merged_cell = self.get_merged_cell_top_left(worksheet, cell_address)
                    if merged_cell:
                        cell_address = merged_cell
                    
                    cell = worksheet[cell_address]
                    row_values.append(cell.value)
                
                if len(row_values) == 1:
                    values.append(row_values[0])
                else:
                    values.extend(row_values)
            
            return values
        except Exception as e:
            raise ValueError(f"Invalid cell range format: {cell_range}") from e
    
    def write_cell_range(self, worksheet: Worksheet, cell_range: str, values: List[Any]) -> None:
        """Write values to cell range.
        
        Args:
            worksheet: Worksheet object
            cell_range: Cell range like "A1:B10" or "C6"
            values: List of values to write
        
        Raises:
            ValueError: Invalid cell range format
            IndexError: Values list size doesn't match cell range
        """
        try:
            start_col, start_row, end_col, end_row = CellUtils.parse_range(cell_range)
            
            start_col_idx = column_index_from_string(start_col)
            end_col_idx = column_index_from_string(end_col)
            
            total_cells = (end_row - start_row + 1) * (end_col_idx - start_col_idx + 1)
            
            if len(values) != total_cells:
                raise IndexError(
                    f"Values list size ({len(values)}) doesn't match cell range size ({total_cells})"
                )
            
            value_idx = 0
            for row in range(start_row, end_row + 1):
                for col_idx in range(start_col_idx, end_col_idx + 1):
                    col_letter = get_column_letter(col_idx)
                    cell_address = f"{col_letter}{row}"
                    
                    merged_cell = self._get_merged_cell_top_left(worksheet, cell_address)
                    if merged_cell:
                        cell_address = merged_cell
                    
                    cell = worksheet[cell_address]
                    cell_value = values[value_idx]
                    cell.value = cell_value
                    
                    if isinstance(cell_value, str) and "\n" in cell_value:
                        if cell.alignment:
                            cell.alignment = Alignment(wrap_text=True, vertical=cell.alignment.vertical, horizontal=cell.alignment.horizontal)
                        else:
                            cell.alignment = Alignment(wrap_text=True)
                    
                    value_idx += 1
        except Exception as e:
            raise ValueError(f"Failed to write cell range: {cell_range}") from e
    
    def is_merged_cell(self, worksheet: Worksheet, cell_address: str) -> bool:
        """Check if cell is part of merged range.
        
        Args:
            worksheet: Worksheet object
            cell_address: Cell address like "A1"
        
        Returns:
            True if cell is merged, False otherwise
        """
        return self._get_merged_cell_top_left(worksheet, cell_address) is not None
    
    def get_merged_cell_top_left(self, worksheet: Worksheet, cell_address: str) -> Optional[str]:
        """Get top-left cell of merged range.
        
        Args:
            worksheet: Worksheet object
            cell_address: Cell address like "A1"
        
        Returns:
            Top-left cell address if merged, None otherwise
        """
        col, row = CellUtils.parse_cell(cell_address)
        col_idx = column_index_from_string(col)
        
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and \
               merged_range.min_col <= col_idx <= merged_range.max_col:
                top_left_col = get_column_letter(merged_range.min_col)
                top_left_row = merged_range.min_row
                return f"{top_left_col}{top_left_row}"
        
        return None
    
    def save_workbook(self, workbook: Workbook, file_path: str) -> None:
        """Save workbook to file.
        
        Args:
            workbook: Workbook object
            file_path: Path to save file
        
        Raises:
            PermissionError: No write permission
            IOError: Failed to save file
        """
        try:
            path = Path(file_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(file_path)
        except PermissionError:
            raise PermissionError(f"No write permission for file: {file_path}")
        except Exception as e:
            raise IOError(f"Failed to save workbook: {e}") from e
    
    def _get_merged_cell_top_left(self, worksheet: Worksheet, cell_address: str) -> Optional[str]:
        """Internal method to get merged cell top-left (alias for get_merged_cell_top_left)."""
        return self.get_merged_cell_top_left(worksheet, cell_address)

